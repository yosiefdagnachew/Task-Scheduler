"""Export schedules to various formats (CSV, ICS, PDF, Excel)."""

import csv
from datetime import datetime
from typing import List, Optional
from pathlib import Path
import pytz
from ics import Calendar, Event
from .models import Schedule, Assignment, TaskType
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from html import escape
import re


def export_to_csv(schedule: Schedule, file_path: str):
    """Export schedule to CSV file."""
    # Ensure output directory exists
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Date', 'Task', 'Role', 'Assignee', 'Week Start (SysAid)'])
        
        # Write one row per assignment with Task and Role columns for dynamic tasks
        def _tt_value(t):
            return t if isinstance(t, str) else t.value

        def _wrap_csv_text(val: str, width: int = 40) -> str:
            """Soft-wrap long values at spaces/hyphens to help viewers render without overflow."""
            if val is None:
                return ''
            s = str(val)
            if len(s) <= width:
                return s
            tokens = re.split(r'(\s+|-)', s)
            lines = []
            cur = ''
            for tok in tokens:
                if len(cur) + len(tok) > width and cur:
                    lines.append(cur.rstrip())
                    cur = tok
                else:
                    cur += tok
            if cur:
                lines.append(cur.rstrip())
            return "\n".join(lines)

        sorted_assignments = sorted(schedule.assignments, key=lambda a: (a.date, _tt_value(a.task_type)))
        for assignment in sorted_assignments:
            week_start_str = assignment.week_start.isoformat() if assignment.week_start else ""
            assignee_display = assignment.assignee.name
            role = assignment.custom_task_shift or assignment.shift_label or ''
            if role and assignment.shift_label and role not in assignee_display:
                # keep assignee name separate
                pass
            writer.writerow([
                assignment.date.isoformat(),
                _wrap_csv_text(_tt_value(assignment.task_type)),
                _wrap_csv_text(role),
                assignee_display,
                week_start_str
            ])


def export_to_ics(schedule: Schedule, file_path: str, timezone: str = "Africa/Addis_Ababa"):
    """Export schedule to ICS calendar file."""
    tz = pytz.timezone(timezone)
    cal = Calendar()
    
    def _tt_value(t):
        return t if isinstance(t, str) else t.value

    def _is_task_eq(t, enum_val):
        return _tt_value(t) == enum_val.value

    # Group assignments by date and task type for better calendar entries
    assignments_by_date = {}
    for assignment in schedule.assignments:
        key = (assignment.date, assignment.task_type, assignment.shift_label)
        if key not in assignments_by_date:
            assignments_by_date[key] = []
        assignments_by_date[key].append(assignment)
    
    # Create events
    for (assignment_date, task_type, shift_label), assignments in assignments_by_date.items():
        assignee_names = ", ".join(a.assignee.name for a in assignments)
        
        # Set times based on task type
        from datetime import time as time_class
        if _is_task_eq(task_type, TaskType.ATM_MORNING):
            # Sunday special: up to 09:00
            if shift_label and "09:00" in shift_label:
                start_time = datetime.combine(assignment_date, time_class(9, 0))
                end_time = datetime.combine(assignment_date, time_class(12, 0))
            else:
                start_time = datetime.combine(assignment_date, time_class(7, 30))
                end_time = datetime.combine(assignment_date, time_class(8, 30))
            title = f"ATM Morning Report - {assignee_names}" + (f" ({shift_label})" if shift_label else "")
        elif _is_task_eq(task_type, TaskType.ATM_MIDNIGHT):
            if shift_label and "06:00" in shift_label:
                start_time = datetime.combine(assignment_date, time_class(6, 0))
                end_time = datetime.combine(assignment_date, time_class(9, 0))
            elif shift_label and "11:00" in shift_label:
                start_time = datetime.combine(assignment_date, time_class(11, 0))
                end_time = datetime.combine(assignment_date, time_class(14, 0))
            elif shift_label and "16:00" in shift_label:
                start_time = datetime.combine(assignment_date, time_class(16, 0))
                end_time = datetime.combine(assignment_date, time_class(22, 0))
            elif shift_label and "09:00" in shift_label:
                start_time = datetime.combine(assignment_date, time_class(9, 0))
                end_time = datetime.combine(assignment_date, time_class(16, 0))
            else:
                start_time = datetime.combine(assignment_date, time_class(13, 0))
                end_time = datetime.combine(assignment_date, time_class(22, 0))
            title = f"ATM Mid-day/Night Report - {assignee_names}" + (f" ({shift_label})" if shift_label else "")
        elif _is_task_eq(task_type, TaskType.SYSAID_MAKER):
            start_time = datetime.combine(assignment_date, time_class(9, 0))
            end_time = datetime.combine(assignment_date, time_class(17, 0))
            title = f"SysAid Maker - {assignee_names}" + (f" ({shift_label})" if shift_label else "")
        elif _is_task_eq(task_type, TaskType.SYSAID_CHECKER):
            start_time = datetime.combine(assignment_date, time_class(9, 0))
            end_time = datetime.combine(assignment_date, time_class(17, 0))
            title = f"SysAid Checker - {assignee_names}" + (f" ({shift_label})" if shift_label else "")
        else:
            start_time = datetime.combine(assignment_date, time_class(9, 0))
            end_time = datetime.combine(assignment_date, time_class(17, 0))
            title = f"{_tt_value(task_type)} - {assignee_names}"
        
        # Localize to timezone
        start_time = tz.localize(start_time)
        end_time = tz.localize(end_time)
        
        event = Event()
        event.name = title
        event.begin = start_time
        event.end = end_time
        event.description = f"Assignee(s): {assignee_names}\nTask: {_tt_value(task_type)}"
        
        cal.events.add(event)
    
    # Write to file
    with open(file_path, 'w', encoding='utf-8') as f:
        f.writelines(cal)


def export_audit_log(audit_log: str, file_path: str):
    """Export audit log to text file."""
    with open(file_path, 'w', encoding='utf-8') as f:
        f.write("SCHEDULING AUDIT LOG\n")
        f.write("=" * 50 + "\n\n")
        f.write(audit_log)


def export_to_xlsx(schedule: Schedule, file_path: str):
    """Export schedule to XLSX with vertical layout (dates as rows)."""
    # Ensure output directory exists
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # If schedule contains only default task types, keep calendar columns; otherwise export vertical rows (Date, Task, Role, Assignee)
    def _tt_value(t):
        return t if isinstance(t, str) else t.value

    task_types = { _tt_value(a.task_type) for a in schedule.assignments }
    default_tasks = {TaskType.ATM_MORNING.value, TaskType.ATM_MIDNIGHT.value, TaskType.SYSAID_MAKER.value, TaskType.SYSAID_CHECKER.value}

    if task_types and task_types.issubset(default_tasks):
        headers = [
            "Date",
            "ATM Morning (07:30-08:30)",
            "ATM Mid/Night",
            "SysAid Maker",
            "SysAid Checker"
        ]
        ws.append(headers)

        by_date = {}
        for a in schedule.assignments:
            d = a.date
            by_date.setdefault(d, []).append(a)

        for day in sorted(by_date.keys()):
            row = [day.isoformat(), "", "", "", ""]
            aggregates: dict[str, list] = {}
            for a in by_date[day]:
                display = a.assignee.name
                if a.shift_label:
                    display = f"{display} ({a.shift_label})"
                aggregates.setdefault(_tt_value(a.task_type), []).append(display)

            row[1] = "\n".join(aggregates.get(TaskType.ATM_MORNING.value, []))
            row[2] = "\n".join(aggregates.get(TaskType.ATM_MIDNIGHT.value, []))
            row[3] = "\n".join(aggregates.get(TaskType.SYSAID_MAKER.value, []))
            row[4] = "\n".join(aggregates.get(TaskType.SYSAID_CHECKER.value, []))
            ws.append(row)
    else:
        # Vertical layout for dynamic/custom tasks
        headers = ["Date", "Task", "Role", "Assignee", "Week Start"]
        ws.append(headers)
        for a in sorted(schedule.assignments, key=lambda x: (x.date, _tt_value(x.task_type))):
            role = a.custom_task_shift or a.shift_label or ''
            week = a.week_start.isoformat() if a.week_start else ''
            assignee = a.assignee.name
            ws.append([a.date.isoformat(), _tt_value(a.task_type), role, assignee, week])

    # Fit columns and wrap text to prevent overlap
    if task_types and task_types.issubset(default_tasks):
        weights = [0.18, 0.205, 0.205, 0.205, 0.205]  # Date, Morning, Mid/Night, Maker, Checker
    else:
        weights = [0.16, 0.34, 0.26, 0.16, 0.08]      # Date, Task, Role, Assignee, Week
    total_chars = 110
    for idx, w in enumerate(weights, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(10, round(total_chars * w))
    # Wrap header and body cells and align to top for readability
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(file_path)


def export_to_excel(schedule: Schedule, file_path: str):
    """Export schedule to Excel format (using openpyxl, same as XLSX)."""
    # Excel format is essentially XLSX, so reuse the XLSX export
    export_to_xlsx(schedule, file_path)


def export_to_pdf(schedule: Schedule, file_path: str):
    """Export schedule to PDF format."""
    # Ensure output directory exists
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=1  # Center
    )
    title = Paragraph(f"Schedule: {schedule.start_date} to {schedule.end_date}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Group assignments by date
    def _tt_value(t):
        return t if isinstance(t, str) else t.value

    by_date = {}
    for a in schedule.assignments:
        d = a.date
        by_date.setdefault(d, []).append(a)
    
    # Table data: adapt layout for default or dynamic tasks
    task_types = { _tt_value(a.task_type) for a in schedule.assignments }
    default_tasks = {TaskType.ATM_MORNING.value, TaskType.ATM_MIDNIGHT.value, TaskType.SYSAID_MAKER.value, TaskType.SYSAID_CHECKER.value}
    is_default_layout = bool(task_types) and task_types.issubset(default_tasks)

    # Paragraph style for table cells with wrapping
    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['BodyText'],
        fontSize=8,
        leading=10,
        wordWrap='CJK',  # robust wrapping, including long tokens
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['BodyText'],
        fontSize=9,
        leading=11,
        wordWrap='CJK',
    )

    def p_multiline(val: str) -> Paragraph:
        """Create a wrapped Paragraph from text, preserving newlines."""
        if val is None:
            val = ""
        parts = str(val).split('\n')
        safe = '<br/>'.join(escape(part) for part in parts)
        return Paragraph(safe or '-', cell_style)
    def p_header(val: str) -> Paragraph:
        return Paragraph(escape(str(val)) if val is not None else '-', header_style)

    if is_default_layout:
        data = [[
            p_header('Date'),
            p_header('ATM Morning'),
            p_header('ATM Mid/Night'),
            p_header('SysAid Maker'),
            p_header('SysAid Checker'),
        ]]
        for day in sorted(by_date.keys()):
            row = [day.strftime('%Y-%m-%d (%A)'), '', '', '', '']
            aggregates: dict[str, list] = {}
            for a in by_date[day]:
                display = a.assignee.name
                if a.shift_label:
                    display = f"{display} ({a.shift_label})"
                aggregates.setdefault(_tt_value(a.task_type), []).append(display)

            row[1] = p_multiline('\n'.join(aggregates.get(TaskType.ATM_MORNING.value, [])) or '-')
            row[2] = p_multiline('\n'.join(aggregates.get(TaskType.ATM_MIDNIGHT.value, [])) or '-')
            row[3] = p_multiline('\n'.join(aggregates.get(TaskType.SYSAID_MAKER.value, [])) or '-')
            row[4] = p_multiline('\n'.join(aggregates.get(TaskType.SYSAID_CHECKER.value, [])) or '-')
            # Convert date to Paragraph for consistency and wrapping if needed
            row[0] = p_multiline(row[0])
            data.append(row)
    else:
        # Vertical rows for dynamic/custom tasks
        data = [[
            p_header('Date'),
            p_header('Task'),
            p_header('Role'),
            p_header('Assignee'),
            p_header('Week Start'),
        ]]
        for day in sorted(by_date.keys()):
            for a in by_date[day]:
                display = a.assignee.name
                role = a.custom_task_shift or a.shift_label or ''
                week = a.week_start.strftime('%Y-%m-%d') if a.week_start else ''
                data.append([
                    p_multiline(day.strftime('%Y-%m-%d (%A)')),
                    p_multiline(_tt_value(a.task_type)),
                    p_multiline(role or '-'),
                    p_multiline(display),
                    p_multiline(week)
                ])
    
    # Create table
    # Compute column widths to fit available page width
    available_width = doc.width
    if is_default_layout:
        # Date, Morning, Mid/Night, Maker, Checker
        weights = [0.18, 0.205, 0.205, 0.205, 0.205]
    else:
        # Date, Task, Role, Assignee, Week
        weights = [0.16, 0.34, 0.26, 0.16, 0.08]
    col_widths = [available_width * w for w in weights]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    
    elements.append(table)
    doc.build(elements)


def export_fairness_to_pdf(fairness_data: List[dict], file_path: str, columns: Optional[List[str]] = None):
    """Export fairness tracking data to PDF.

    If `columns` is provided, render a dynamic header using those task names; otherwise
    fall back to the default fixed columns for built-in tasks.
    """
    # Ensure output directory exists
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(file_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=1  # Center
    )
    title = Paragraph("Fairness Tracking Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Styles for header and cells so long text wraps and fits page
    header_style = ParagraphStyle(
        'FairHeader',
        parent=styles['BodyText'],
        fontSize=9,
        leading=11,
        wordWrap='CJK',
    )
    cell_style = ParagraphStyle(
        'FairCell',
        parent=styles['BodyText'],
        fontSize=9,
        leading=11,
        wordWrap='CJK',
    )

    def ph(val: str) -> Paragraph:
        return Paragraph(escape(str(val)) if val is not None else '-', header_style)

    def p(val: str) -> Paragraph:
        return Paragraph(escape(str(val)) if val is not None else '-', cell_style)

    # Table data
    if columns is None or len(columns) == 0:
        # Default built-in tasks
        header = [ph('Member'), ph('ATM Morning'), ph('ATM Mid/Night'), ph('SysAid Maker'), ph('SysAid Checker'), ph('Total')]
        data = [header]
        for member in fairness_data:
            row = [
                p(member.get('member_name', member.get('member_id', 'Unknown'))),
                p(str(member.get('counts', {}).get('ATM_MORNING', 0))),
                p(str(member.get('counts', {}).get('ATM_MIDNIGHT', 0))),
                p(str(member.get('counts', {}).get('SYSAID_MAKER', 0))),
                p(str(member.get('counts', {}).get('SYSAID_CHECKER', 0))),
                p(str(member.get('total', 0)))
            ]
            data.append(row)
        # Compute widths to fit page width
        available_width = doc.width
        mem_w = 0.32
        tot_w = 0.10
        per_task = (1.0 - mem_w - tot_w) / 4.0
        weights = [mem_w] + [per_task, per_task, per_task, per_task] + [tot_w]
        col_widths = [available_width * w for w in weights]
    else:
        # Dynamic tasks based on provided columns
        header = [ph('Member')] + [ph(c) for c in list(columns)] + [ph('Total')]
        data = [header]
        for member in fairness_data:
            counts = member.get('counts', {})
            row = [p(member.get('member_name', member.get('member_id', 'Unknown')))]
            for c in columns:
                row.append(p(str(counts.get(c, 0))))
            row.append(p(str(member.get('total', 0))))
            data.append(row)
        # Compute col widths based on available page width
        available_width = doc.width
        task_count = max(1, len(columns))
        mem_w = 0.34 if task_count > 4 else 0.32
        tot_w = 0.10
        per_task = (1.0 - mem_w - tot_w) / task_count
        weights = [mem_w] + [per_task for _ in columns] + [tot_w]
        col_widths = [available_width * w for w in weights]
    
    # Create table
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    
    elements.append(table)
    doc.build(elements)

